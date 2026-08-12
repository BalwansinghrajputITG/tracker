"""
Excel export (hr.md §39).

openpyxl in write_only mode: rows stream to disk as they are added rather than
being held as cell objects, which is what keeps a 5000-row export from costing
tens of megabytes of resident memory per concurrent request.

CSV and Excel cover all nine §39 report types. PDF is reserved for the three
letter artifacts (offer, experience, relieving) where layout is the point —
good PDF tables need system libraries, and a report is more useful in a
spreadsheet than as a picture of one.
"""

from __future__ import annotations

import io
from datetime import date, datetime
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


def _cell(value):
    """Coerce a value into something openpyxl can write.

    Mirrors utils/export._cell, including the leading-character guard: a string
    starting =, +, - or @ is evaluated as a formula by Excel and Sheets.
    """
    if value is None:
        return ""
    if isinstance(value, (int, float, bool, datetime, date)):
        return value
    text = str(value)
    if text[:1] in ("=", "+", "-", "@"):
        return "'" + text
    return text


def build_workbook(
    sheets: Sequence[tuple[str, Sequence[str], Iterable[Sequence]]],
) -> bytes:
    """Build an .xlsx from (sheet_name, headers, rows) tuples.

    Multiple sheets in one file because an HR report is usually a summary plus
    its detail, and two downloads to answer one question is a worse answer.
    """
    wb = Workbook(write_only=True)

    for name, headers, rows in sheets:
        # Excel rejects >31 chars and the characters below in a sheet name; it
        # fails at open time, not at write time, so sanitise here.
        safe = str(name)[:31]
        for ch in "[]:*?/\\":
            safe = safe.replace(ch, "-")
        ws = wb.create_sheet(safe)

        widths = [max(12, min(40, len(str(h)) + 4)) for h in headers]
        ws.column_dimensions_holder = None  # write_only: set via column_dimensions below
        for index, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(index)].width = width

        header_cells = []
        from openpyxl.cell import WriteOnlyCell
        for h in headers:
            c = WriteOnlyCell(ws, value=str(h))
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(vertical="center")
            header_cells.append(c)
        ws.append(header_cells)

        for row in rows:
            ws.append([_cell(v) for v in row])

    if not wb.sheetnames:
        wb.create_sheet("Empty")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def xlsx_headers(filename: str) -> dict:
    return {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "X-Content-Type-Options": "nosniff",
        "Cache-Control": "private, no-store",
    }
